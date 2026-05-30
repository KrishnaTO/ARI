"""
Second pass: resolve the NO-MATCH rows in the Athena_Match sheet by searching
the OHDSI Athena API with the *term name* instead of the SNOMED code.

These rows failed code matching because their SNOMED code isn't in Athena's
vocabulary snapshot. A name search can still find the concept, but it will have
a DIFFERENT SNOMED code, so every resolution is flagged and the code difference
recorded in a new `nameMatchNote` column. Rows with no exact SNOMED name match
are left as NO MATCH with their top candidates listed for manual review.

Auth via env vars (same as athena_lookup.py):
    ATHENA_TOKEN, ATHENA_COOKIE

Usage:  python athena_namematch.py
"""

import os
import re
import sys
import time

import openpyxl

from athena_lookup import (
    WORKBOOK, OUTPUT_SHEET, make_session, search, detail, relationships,
    fmt_synonyms, fmt_relationships,
)

NOTE_COL = 20  # new column appended after the existing 19


def norm(s):
    s = (s or "").lower().strip()
    s = re.sub(r"\s*\([^)]*\)\s*$", "", s)   # drop a trailing "(disorder)" etc.
    s = re.sub(r"\s+", " ", s)
    return s


def pick(term, hits):
    """Return (hit, how) for the best exact-name SNOMED match, else (None, None)."""
    nt = norm(term)
    exact = [h for h in hits if norm(h.get("name")) == nt]
    snomed_std = [h for h in exact
                  if h.get("vocabulary") == "SNOMED" and h.get("standardConcept") == "Standard"]
    snomed_any = [h for h in exact if h.get("vocabulary") == "SNOMED"]
    if snomed_std:
        return snomed_std[0], "exact-name SNOMED Standard"
    if snomed_any:
        return snomed_any[0], "exact-name SNOMED"
    return None, None


def candidate_note(hits, limit=5):
    parts = []
    for h in hits[:limit]:
        parts.append(f"{h.get('code')}|{h.get('vocabulary')}|"
                     f"{h.get('standardConcept')}|{h.get('name')}")
    return "candidates: " + " ; ".join(parts) if parts else "no search results"


def run():
    session = make_session()
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[OUTPUT_SHEET]

    if ws.cell(row=1, column=NOTE_COL).value != "nameMatchNote":
        ws.cell(row=1, column=NOTE_COL, value="nameMatchNote")

    targets = [r for r in range(2, ws.max_row + 1)
               if ws.cell(row=r, column=3).value == "NO MATCH"]
    print(f"{len(targets)} NO-MATCH rows to retry by name\n")

    resolved = review = 0
    for n, r in enumerate(targets, 1):
        term = ws.cell(row=r, column=1).value
        input_code = ws.cell(row=r, column=2).value
        try:
            hits = search(session, term)
            hit, how = pick(term, hits)
            if hit:
                omop_id = hit["id"]
                det = detail(session, omop_id)
                rel = relationships(session, omop_id)
                vals = {
                    3: "OK (name match)", 4: "term-name", 5: omop_id,
                    6: det.get("name"), 7: det.get("conceptCode"),
                    8: det.get("vocabularyId"), 9: det.get("domainId"),
                    10: det.get("conceptClassId"), 11: det.get("standardConcept"),
                    12: det.get("invalidReason"),
                    13: str(det.get("validStart") or "")[:10],
                    14: str(det.get("validEnd") or "")[:10],
                    15: fmt_synonyms(det), 16: rel.get("count"),
                    17: fmt_relationships(rel), 18: det.get("vocabularyName"),
                    19: det.get("vocabularyVersion"),
                    NOTE_COL: (f"{how}; Athena code {det.get('conceptCode')} "
                               f"differs from input {input_code}"),
                }
                for col, v in vals.items():
                    ws.cell(row=r, column=col, value=v)
                resolved += 1
                status = f"RESOLVED -> {det.get('conceptCode')}"
            else:
                ws.cell(row=r, column=NOTE_COL, value=candidate_note(hits))
                review += 1
                status = "still no exact SNOMED name match"
        except Exception as e:  # noqa: BLE001
            ws.cell(row=r, column=NOTE_COL, value=f"ERROR: {type(e).__name__}")
            status = f"ERROR {type(e).__name__}"

        print(f"[{n}/{len(targets)}] {status:35} | {input_code}  {term}")
        time.sleep(0.25)

    ws.column_dimensions["T"].width = 70
    wb.save(WORKBOOK)
    print(f"\nDone. Resolved by name: {resolved} | Left for review: {review}")


if __name__ == "__main__":
    run()
