#!/usr/bin/env python3
import openpyxl, os, re, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

BASE="/sessions/clever-zen-ritchie/mnt/data"
OUTDIR=os.path.join(BASE,"ARI-Linked-Database")
OUT=os.path.join(OUTDIR,"ARI-Linked-Disease-Database.xlsx")
resolved=json.load(open(os.path.join(OUTDIR,"_resolved.json")))
crosswalk=json.load(open(os.path.join(OUTDIR,"_crosswalk.json")))

def load(f,sheet):
    wb=openpyxl.load_workbook(os.path.join(BASE,f),read_only=True,data_only=True);ws=wb[sheet]
    data=list(ws.iter_rows(values_only=True));wb.close()
    hdr=[(str(c).strip() if c is not None else c) for c in data[0]]
    return [dict(zip(hdr,r)) for r in data[1:] if any(v is not None for v in r)]
def s(v):
    if v is None: return None
    v=str(v).strip(); return v if v else None
def nsn(v):
    v=s(v)
    if not v: return None
    m=re.match(r'^\d{6,}$',v.replace(" ",""));return m.group(0) if m else None

# SNOMED -> ARI_ID , DOID -> ARI_ID
sno2id={}; doid2id={}
for r in resolved:
    for c in r['SNOMED']: sno2id[c]=r['ARI_ID']
    for d in r['DOID']: doid2id[d]=r['ARI_ID']
name_by_id={r['ARI_ID']:r['Primary_Name'] for r in resolved}

# ---- styling helpers ----
HDR_FILL=PatternFill('solid',fgColor='1F4E78')
HDR_FONT=Font(name='Arial',bold=True,color='FFFFFF',size=10)
CELL_FONT=Font(name='Arial',size=10)
WRAP=Alignment(wrap_text=True,vertical='top')
TOP=Alignment(vertical='top')
thin=Side(style='thin',color='D9D9D9')
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def write_sheet(wb,title,headers,rows,widths=None,wrapcols=None,first=False):
    ws=wb.active if first else wb.create_sheet()
    ws.title=title
    wrapcols=wrapcols or set()
    for j,h in enumerate(headers,1):
        c=ws.cell(1,j,h); c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=Alignment(vertical='center'); c.border=BORDER
    for i,row in enumerate(rows,2):
        for j,val in enumerate(row,1):
            c=ws.cell(i,j,val); c.font=CELL_FONT; c.border=BORDER
            c.alignment=WRAP if (j-1) in wrapcols else TOP
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}{max(2,len(rows)+1)}"
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.row_dimensions[1].height=22
    return ws

wb=Workbook()

# ============ 1. README ============
ws=wb.active; ws.title="README"
ws.column_dimensions['A'].width=26; ws.column_dimensions['B'].width=110
ws['A1']="ARI Linked Disease Database"; ws['A1'].font=Font(name='Arial',bold=True,size=16,color='1F4E78')
import datetime
lines=[
 ("Generated", datetime.date.today().isoformat()),
 ("Purpose","Fully-linked relational database unifying ARI autoimmune-disease data across DOID, SNOMED CT, MESH and OMOP/Athena."),
 ("Scope","Union of all source concepts (ARI curated list + full DOID & SNOMED autoimmune subsets), deduplicated by shared identifiers."),
 ("",""),
 ("IDENTITY MODEL",""),
 ("Concept key","Same SNOMED conceptId = same concept; same DOID CURIE = same concept."),
 ("SNOMED<->DOID merge","Merged ONLY for reciprocal 1:1 matches (mutually unique). 34 such merges."),
 ("Other links","All other SNOMED<->DOID links are kept in Crosswalk_SNOMED_DOID (not merged), because ARI's DOID column and DOID db_xref are many-to-many name/xref lookups."),
 ("ARI_ID","Generated surrogate primary key (ARI-#####). The foreign key joining every sheet."),
 ("DOID normalization","All DOID CURIEs canonicalised to 7-digit form (e.g. doid:4313 -> DOID:0004313)."),
 ("",""),
 ("SHEETS",""),
 ("Diseases","Master concept table (one row per concept). PK = ARI_ID."),
 ("Master_Wide","Denormalized one-row-per-concept view with synonyms & cross-refs collapsed into delimited cells."),
 ("Synonyms","ARI_ID -> synonym (long format). Source-tagged."),
 ("Definitions","ARI_ID -> definition text (DOID / SNOMED), source-tagged."),
 ("CrossReferences","ARI_ID -> external vocab refs from DOID db_xref (MESH, ICD10CM, ICD9CM, NCI, UMLS_CUI, ORDO, GARD, MIM, EFO, KEGG)."),
 ("Crosswalk_SNOMED_DOID","All candidate SNOMED<->DOID links with basis + whether merged. Links SNOMED-anchored & DOID-anchored concepts."),
 ("SNOMED_Details","Full SNOMED CT metadata (FSN, preferred term, semantic tag, definition status, text definition, module, effective time)."),
 ("Athena_OMOP","OMOP/Athena standard-concept mappings (OMOP conceptId, vocabulary, domain, class, validity)."),
 ("",""),
 ("SOURCES",""),
 ("ARI-doid","ARI-doid/ARI-doid.xlsx - ARI curated disease list, DOID lookup by name, SNOMED + OMOP ConceptID."),
 ("ARI-MESH","ARI-MESH_Synonyms/ARI-MESHsynonyms-DOID.xlsx - MESH-sourced synonyms, DOID lookup by name."),
 ("ARI-SNOMED-Lookup","ARI-SNOMED-Athena/ARI_SNOMED_Lookup.xlsx - up-to-date SNOMED data per ARI SNOMED code."),
 ("ARI-Athena","ARI-SNOMED-Athena/ARI_Athena_Matches.xlsx - ARI lookup to Athena/OMOP."),
 ("DOID-all","DOID_autoimmune_diseases/DOID-all.xlsx - all DOID children of 'autoimmune disease' + synonyms, definitions, db_xref."),
 ("SNOMED-all","SNOMED-Athena/SNOMED_Athena_Matches-all_autoimmune_disease.xlsx - all SNOMED autoimmune concepts + Athena matches."),
]
r=3
for k,v in lines:
    a=ws.cell(r,1,k); b=ws.cell(r,2,v)
    a.font=Font(name='Arial',bold=bool(k) and v=='' or k in ('Generated','Purpose','Scope'),size=10)
    if v=='' and k: a.font=Font(name='Arial',bold=True,size=11,color='1F4E78')
    else: a.font=Font(name='Arial',bold=True,size=10)
    b.font=CELL_FONT; b.alignment=WRAP; a.alignment=TOP
    r+=1

# ============ 2. Diseases (master) ============
hdr=["ARI_ID","Primary_Name","Category","ARI_Parent","In_ARI","Semantic_Tag","SNOMED_ID","DOID_CURIE","OMOP_ConceptIDs","Definition","n_Synonyms","n_CrossRefs","Sources"]
rows=[]
for r in resolved:
    bestdef=r['Definitions'][0][0] if r['Definitions'] else None
    rows.append([r['ARI_ID'],r['Primary_Name'],r['Category'],r['ARI_Parent'],r['In_ARI'],r['SemanticTag'],
        "; ".join(r['SNOMED']) or None,"; ".join(r['DOID']) or None,"; ".join(r['OMOP']) or None,
        bestdef,len(r['Synonyms']),len(r['Xrefs']),"; ".join(r['Sources'])])
write_sheet(wb,"Diseases",hdr,rows,
    widths=[11,40,16,32,7,14,16,16,20,60,11,11,40],wrapcols={1,3,9})

# ============ 3. Master_Wide (denormalized) ============
XR_SYS=["MESH","ICD10CM","ICD9CM","NCI","UMLS_CUI","ORDO","GARD","MIM","EFO","KEGG"]
hdr=["ARI_ID","Primary_Name","Category","ARI_Parent","In_ARI","SNOMED_ID","DOID_CURIE","OMOP_ConceptIDs","Semantic_Tag","All_Synonyms"]+XR_SYS+["Other_XRefs","Definition","Sources"]
rows=[]
for r in resolved:
    xr=defaultdict(list); other=[]
    for sysn,val,src in r['Xrefs']:
        base=sysn.split('_US_')[0] if sysn.upper().startswith('SNOMED') else sysn
        key=None
        for k in XR_SYS:
            if base.upper()==k or base.upper().startswith(k): key=k; break
        if base.upper().startswith('SNOMED'): continue  # SNOMED captured separately
        if key: xr[key].append(val)
        else: other.append(f"{sysn}:{val}")
    syns=" | ".join(t for t,_,_ in r['Synonyms'])
    bestdef=r['Definitions'][0][0] if r['Definitions'] else None
    row=[r['ARI_ID'],r['Primary_Name'],r['Category'],r['ARI_Parent'],r['In_ARI'],
         "; ".join(r['SNOMED']) or None,"; ".join(r['DOID']) or None,"; ".join(r['OMOP']) or None,
         r['SemanticTag'],syns or None]
    for k in XR_SYS: row.append("; ".join(xr[k]) or None)
    row+=[" | ".join(other) or None,bestdef,"; ".join(r['Sources'])]
    rows.append(row)
write_sheet(wb,"Master_Wide",hdr,rows,
    widths=[11,38,15,30,7,16,16,18,13,55]+[14]*len(XR_SYS)+[24,55,38],wrapcols={1,3,9,len(hdr)-2})

# ============ 4. Synonyms ============
hdr=["ARI_ID","Primary_Name","Synonym","Synonym_Type","Source"]
rows=[]
for r in resolved:
    for txt,typ,src in r['Synonyms']:
        rows.append([r['ARI_ID'],r['Primary_Name'],txt,typ,src])
write_sheet(wb,"Synonyms",hdr,rows,widths=[11,38,45,28,20],wrapcols={2})

# ============ 5. Definitions ============
hdr=["ARI_ID","Primary_Name","Definition","Def_System","Source_File"]
rows=[]
for r in resolved:
    for txt,sysn,src in r['Definitions']:
        rows.append([r['ARI_ID'],r['Primary_Name'],txt,sysn,src])
write_sheet(wb,"Definitions",hdr,rows,widths=[11,38,90,12,20],wrapcols={2})

# ============ 6. CrossReferences ============
hdr=["ARI_ID","Primary_Name","DOID_CURIE","Xref_System","Xref_ID","Source"]
rows=[]
for r in resolved:
    doid="; ".join(r['DOID'])
    for sysn,val,src in r['Xrefs']:
        rows.append([r['ARI_ID'],r['Primary_Name'],doid or None,sysn,val,src])
write_sheet(wb,"CrossReferences",hdr,rows,widths=[11,38,16,28,18,16])

# ============ 7. Crosswalk_SNOMED_DOID ============
hdr=["SNOMED_ID","DOID_CURIE","Link_Basis","Merged_1to1","SNOMED_ARI_ID","DOID_ARI_ID","SNOMED_Name","DOID_Name"]
rows=[]
for c in crosswalk:
    rows.append([c['SNOMED'],c['DOID'],c['Basis'],c['Merged'],c['SNOMED_ARI_ID'],c['DOID_ARI_ID'],
        name_by_id.get(c['SNOMED_ARI_ID']),name_by_id.get(c['DOID_ARI_ID'])])
write_sheet(wb,"Crosswalk_SNOMED_DOID",hdr,rows,widths=[16,16,22,12,13,13,36,36],wrapcols={6,7})

# ============ 8. SNOMED_Details ============
snolook=load("ARI-SNOMED-Athena/ARI_SNOMED_Lookup.xlsx","ARI-SNOMED Lookup")
sno_all=load("SNOMED-Athena/SNOMED_Athena_Matches-all_autoimmune_disease.xlsx","SNOMED-AutoimmuneDisease")
det={}  # snomed -> dict
for r in snolook:
    sn=nsn(r.get('ARI_SNOMED_ID'))
    if not sn: continue
    det[sn]={'FSN':s(r.get('FSN')),'PreferredTerm':s(r.get('PreferredTerm')),'SemanticTag':s(r.get('SemanticTag')),
        'DefinitionStatus':s(r.get('DefinitionStatus')),'ConceptActive':s(r.get('Concept_Active')),
        'InRelease':s(r.get('In_Release')),'TextDefinition':s(r.get('TextDefinition')),
        'ModuleId':s(r.get('ModuleId')),'EffectiveTime':s(r.get('EffectiveTime')),'Source':'ARI-SNOMED-Lookup'}
for r in sno_all:
    sn=nsn(r.get('conceptId'))
    if not sn: continue
    d=det.setdefault(sn,{'Source':'SNOMED-all'})
    d.setdefault('PreferredTerm',s(r.get('term')))
    if not d.get('FSN'): d['FSN']=None
    d.setdefault('SemanticTag',s(r.get('semanticTag')))
    d.setdefault('DefinitionStatus',s(r.get('definitionStatus')))
    if d.get('ConceptActive') is None: d['ConceptActive']=s(r.get('conceptActive'))
    d.setdefault('TextDefinition',s(r.get('textDefinition')))
    if 'Source' in d and d['Source']=='ARI-SNOMED-Lookup': d['Source']='ARI-SNOMED-Lookup; SNOMED-all'
hdr=["ARI_ID","SNOMED_ID","FSN","PreferredTerm","SemanticTag","DefinitionStatus","ConceptActive","In_Release","TextDefinition","ModuleId","EffectiveTime","Source"]
rows=[]
for sn,d in sorted(det.items()):
    rows.append([sno2id.get(sn),sn,d.get('FSN'),d.get('PreferredTerm'),d.get('SemanticTag'),
        d.get('DefinitionStatus'),d.get('ConceptActive'),d.get('InRelease'),d.get('TextDefinition'),
        d.get('ModuleId'),d.get('EffectiveTime'),d.get('Source')])
write_sheet(wb,"SNOMED_Details",hdr,rows,widths=[11,16,45,38,14,16,13,11,70,20,14,22],wrapcols={2,3,8})

# ============ 9. Athena_OMOP ============
ari_ath=load("ARI-SNOMED-Athena/ARI_Athena_Matches.xlsx","Matches")
sno_ath=load("SNOMED-Athena/SNOMED_Athena_Matches-all_autoimmune_disease.xlsx","Athena_Match")
om={}  # (snomed,omop) -> dict
for r in sno_ath:
    sn=nsn(r.get('inputConceptId(SNOMED)')); omop=s(r.get('athenaConceptId(OMOP)'))
    if not sn: continue
    om[(sn,omop)]={'AthenaName':s(r.get('athenaName')),'vocabularyId':s(r.get('vocabularyId')),
        'domainId':s(r.get('domainId')),'conceptClassId':s(r.get('conceptClassId')),
        'standardConcept':s(r.get('standardConcept')),'invalidReason':s(r.get('invalidReason')),
        'validStart':s(r.get('validStart')),'validEnd':s(r.get('validEnd')),
        'matchStatus':s(r.get('matchStatus')),'Source':'SNOMED-Athena'}
for r in ari_ath:
    sn=nsn(r.get('LoD_SNOMED')) or nsn(r.get('Athena_SNOMED')); omop=s(r.get('Athena_OMOP'))
    if not sn: continue
    key=(sn,omop)
    if key not in om:
        om[key]={'AthenaName':s(r.get('Athena_OMOP_Name')),'vocabularyId':s(r.get('Athena_vocabularyId')),
            'matchStatus':s(r.get('Athena_matchStatus')),'Source':'ARI-Athena'}
    else:
        om[key]['Source']='ARI-Athena; SNOMED-Athena'
hdr=["ARI_ID","SNOMED_ID","OMOP_ConceptID","AthenaName","vocabularyId","domainId","conceptClassId","standardConcept","invalidReason","validStart","validEnd","matchStatus","Source"]
rows=[]
for (sn,omop),d in sorted(om.items(),key=lambda kv:(kv[0][0],kv[0][1] or '')):
    rows.append([sno2id.get(sn),sn,omop,d.get('AthenaName'),d.get('vocabularyId'),d.get('domainId'),
        d.get('conceptClassId'),d.get('standardConcept'),d.get('invalidReason'),d.get('validStart'),
        d.get('validEnd'),d.get('matchStatus'),d.get('Source')])
write_sheet(wb,"Athena_OMOP",hdr,rows,widths=[11,16,15,40,13,13,15,15,13,12,12,12,22],wrapcols={3})

wb.save(OUT)
print("Saved:",OUT)
print("Sheets:",wb.sheetnames)
for sn in wb.sheetnames:
    print(f"  {sn}: {wb[sn].max_row-1} rows")
