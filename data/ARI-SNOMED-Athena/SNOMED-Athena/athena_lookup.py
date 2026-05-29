"""
Look up SNOMED terms on the OHDSI Athena API and build a consolidated table
containing the full matching entry: OMOP concept id, SNOMED code, metadata,
synonyms, and term connections (relationships).

Auth: Athena's API requires a session token captured from a logged-in browser
(the `Athena-Auth-Token` header + session cookies). These are passed in via
environment variables so the script can be re-run without editing it.

    ATHENA_TOKEN   - value of the Athena-Auth-Token header
    ATHENA_COOKIE  - value of the Cookie header (athena_js_session=...; JSESSIONID=...)

Usage:
    python athena_lookup.py --limit 15      # first 15 rows (sample)
    python athena_lookup.py                  # all rows
"""

import argparse
import os
import sys
import time

import openpyxl
import requests

API = "https://athena.ohdsi.org/api/v1"
WORKBOOK = os.path.join(os.path.dirname(__file__), "autoimmune_disease_combined_terms.xlsx")
SOURCE_SHEET = "SNOMED-AutoimmuneDisease"
OUTPUT_SHEET = "Athena_Match"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36 Edg/148.0.0.0")


def make_session():
    token = os.environ.get("ATHENA_TOKEN")
    cookie = os.environ.get("ATHENA_COOKIE")
    if not token or not cookie:
        sys.exit("ERROR: set ATHENA_TOKEN and ATHENA_COOKIE environment variables.")
    s = requests.Session()
    s.headers.update({
        "Accept": "application/json",
        "Athena-Auth-Token": token,
        "Cookie": cookie,
        "User-Agent": UA,
        "Referer": "https://athena.ohdsi.org/search-terms/start",
    })
    return s


def get_json(session, path, params=None, retries=4):
    """GET with retry/backoff. Raises the last error if all attempts fail."""
    last = None
    for attempt in range(retries):
        try:
            r = session.get(f"{API}/{path}", params=params, timeout=30)
            if r.status_code in (401, 403):
                # auth problem - retrying won't help, surface immediately
                r.raise_for_status()
            r.raise_for_status()
            return r.json()
        except requests.HTTPError as e:
            if e.response is not None and e.response.status_code in (401, 403):
                raise
            last = e
        except (requests.ConnectionError, requests.Timeout, ValueError) as e:
            last = e
        time.sleep(1.5 * (attempt + 1))
    raise last


def search(session, query, page_size=20):
    return get_json(session, "concepts",
                    {"query": query, "pageSize": page_size}).get("content", [])


def detail(session, omop_id):
    return get_json(session, f"concepts/{omop_id}")


def relationships(session, omop_id):
    return get_json(session, f"concepts/{omop_id}/relationships")


def find_match(session, concept_id, term):
    """Return the Athena search hit for this SNOMED concept, or None.

    Strategy: search by the SNOMED code, keep SNOMED-vocabulary hits whose code
    matches exactly. Fall back to a term-name search if the code search misses.
    """
    code = str(concept_id)
    for query in (code, term):
        hits = search(session, query)
        exact = [h for h in hits
                 if h.get("vocabulary") == "SNOMED" and str(h.get("code")) == code]
        if exact:
            # Prefer a Standard concept if more than one
            exact.sort(key=lambda h: 0 if h.get("standardConcept") == "Standard" else 1)
            return exact[0], ("code" if query == code else "term")
    return None, None


def fmt_synonyms(det):
    out = []
    for syn in det.get("synonyms", []) or []:
        name = (syn.get("synonymName") or "").strip()
        lang = (syn.get("langName") or "").strip()
        out.append(f"{name} [{lang}]" if lang else name)
    return " ; ".join(out)


def fmt_relationships(rel):
    groups = []
    for grp in rel.get("items", []) or []:
        gname = grp.get("relationshipName", "")
        targets = []
        for t in grp.get("relationships", []) or []:
            targets.append(f"{t.get('targetConceptId')}|"
                           f"{t.get('targetConceptName')} ({t.get('targetVocabularyId')})")
        groups.append(f"{gname}: " + "; ".join(targets))
    return " || ".join(groups)


HEADERS = [
    "inputTerm", "inputConceptId(SNOMED)", "matchStatus", "matchedBy",
    "athenaConceptId(OMOP)", "athenaName", "conceptCode", "vocabularyId",
    "domainId", "conceptClassId", "standardConcept", "invalidReason",
    "validStart", "validEnd", "synonyms", "relationshipCount",
    "termConnections", "vocabularyName", "vocabularyVersion",
]


def process(limit=None):
    session = make_session()
    wb = openpyxl.load_workbook(WORKBOOK)
    src = wb[SOURCE_SHEET]
    rows = list(src.iter_rows(min_row=2, values_only=True))
    if limit:
        rows = rows[:limit]

    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    out = wb.create_sheet(OUTPUT_SHEET)
    out.append(HEADERS)

    counts = {"OK": 0, "NO MATCH": 0, "ERROR": 0}
    for i, row in enumerate(rows, 1):
        concept_id, term = row[0], row[1]
        try:
            hit, by = find_match(session, concept_id, term)
            if not hit:
                out.append([term, concept_id, "NO MATCH"] + [""] * (len(HEADERS) - 3))
                counts["NO MATCH"] += 1
                status = "NO MATCH"
            else:
                omop_id = hit["id"]
                det = detail(session, omop_id)
                rel = relationships(session, omop_id)
                out.append([
                    term, concept_id, "OK", by, omop_id,
                    det.get("name"), det.get("conceptCode"), det.get("vocabularyId"),
                    det.get("domainId"), det.get("conceptClassId"),
                    det.get("standardConcept"), det.get("invalidReason"),
                    str(det.get("validStart") or "")[:10],
                    str(det.get("validEnd") or "")[:10],
                    fmt_synonyms(det), rel.get("count"), fmt_relationships(rel),
                    det.get("vocabularyName"), det.get("vocabularyVersion"),
                ])
                counts["OK"] += 1
                status = "OK"
        except requests.HTTPError as e:
            sc = e.response.status_code if e.response is not None else "?"
            if sc in (401, 403):
                print(f"[{i}] AUTH FAILURE (HTTP {sc}) - token/session likely expired. "
                      f"Saving progress and stopping.")
                wb.save(WORKBOOK)
                sys.exit(f"Stopped at row {i} due to auth failure. "
                         f"Paste a fresh token and re-run.")
            out.append([term, concept_id, f"ERROR: HTTP {sc}"] + [""] * (len(HEADERS) - 3))
            counts["ERROR"] += 1
            status = f"ERROR {sc}"
        except Exception as e:  # noqa: BLE001 - record and continue
            out.append([term, concept_id, f"ERROR: {type(e).__name__}"] + [""] * (len(HEADERS) - 3))
            counts["ERROR"] += 1
            status = f"ERROR {type(e).__name__}"

        print(f"[{i}/{len(rows)}] {status:9} {concept_id}  {term}")
        if i % 50 == 0:
            wb.save(WORKBOOK)
            print(f"  ...checkpoint saved at {i} rows  (OK={counts['OK']} "
                  f"NO MATCH={counts['NO MATCH']} ERROR={counts['ERROR']})")
        time.sleep(0.25)

    # widen a few columns for readability
    widths = {"A": 45, "O": 60, "Q": 80, "F": 40}
    for col, w in widths.items():
        out.column_dimensions[col].width = w

    wb.save(WORKBOOK)
    print(f"\nSaved sheet '{OUTPUT_SHEET}' with {len(rows)} term(s) to {WORKBOOK}")
    print(f"Summary: OK={counts['OK']}  NO MATCH={counts['NO MATCH']}  ERROR={counts['ERROR']}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    args = ap.parse_args()
    process(limit=args.limit)
