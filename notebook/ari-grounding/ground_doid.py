"""Ground ALL ARI core diseases to DOID using Gilda lexical matching.

Method mirrors notebook/build_autoimmune-single.py (pyobo's get_grounder uses Gilda
under the hood). Data source is LOCAL: data/2-databases/doid.owl (parsed to
doid_records.json). No online sources.
"""
import json, csv, os, re, html
from gilda import Grounder
from gilda.term import Term
from gilda.process import normalize
import openpyxl

BASE = "/sessions/zen-keen-curie/mnt/ARI"
DOID_RECORDS = "/sessions/zen-keen-curie/mnt/ARI/notebook/ari-grounding/doid_records.json"
CORE = f"{BASE}/data/4-reports/1_Core_ARI_Diseases.xlsx"
OUT = f"{BASE}/notebook/ari-grounding"
os.makedirs(OUT, exist_ok=True)

records = json.load(open(DOID_RECORDS))

# Build Gilda terms from local DOID labels + synonyms
terms = []
for doid_id, rec in records.items():
    num = doid_id.split(":")[1]
    label = rec.get("label", "")
    if label:
        terms.append(Term(normalize(label), label, "DOID", num, label, "name", "doid"))
    for syn in rec.get("exact_syn", []):
        if syn:
            terms.append(Term(normalize(syn), syn, "DOID", num, label or syn, "synonym", "doid"))
    for syn in rec.get("related_syn", []) + rec.get("narrow_syn", []):
        if syn:
            terms.append(Term(normalize(syn), syn, "DOID", num, label or syn, "former_name", "doid"))

grounder = Grounder(terms)
print(f"DOID grounder built from {len(terms)} local terms ({len(records)} classes)")

# Load all core diseases (ARI ID, Preferred Name, Synonyms)
cws = openpyxl.load_workbook(CORE, read_only=True).active
hdr = [c for c in next(cws.iter_rows(min_row=1, max_row=1, values_only=True))]
i_ari, i_name, i_syn = hdr.index("ARI ID"), hdr.index("Preferred Name"), hdr.index("Synonyms")
diseases = [(r[i_ari], r[i_name], r[i_syn] or "") for r in cws.iter_rows(min_row=2, values_only=True)]

def best_match(name, synonyms):
    """Ground preferred name first; fall back to synonyms. Return best ScoredMatch + which text matched."""
    candidates = [name] + [s.strip() for s in synonyms.split(";") if s.strip()]
    best = None; best_via = ""
    for txt in candidates:
        for m in grounder.ground(txt):
            if best is None or m.score > best.score:
                best = m; best_via = txt
        if best and best.score >= 0.9 and best_via == name:
            break  # strong match on preferred name
    return best, best_via

rows = []
matched = 0
for ari, name, syn in diseases:
    m, via = best_match(name, syn)
    if m:
        matched += 1
        rec = records.get("DOID:"+m.term.id, {})
        rows.append([ari, name, f"DOID:{m.term.id}", m.term.entry_name,
                     round(m.score, 3), m.term.status, via if via != name else "(preferred name)",
                     ", ".join(rec.get("snomed", [])[:5]),
                     "Yes" if rec.get("deprecated") else ""])
    else:
        rows.append([ari, name, "", "", "", "", "no match", "", ""])

with open(f"{OUT}/doid_matches_all.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["ARI ID","Preferred Name","DOID","DOID Label","Score","Match Type",
                "Matched Via","DOID SNOMED xref","DOID Obsolete"])
    w.writerows(rows)

print(f"Diseases: {len(diseases)} | DOID matched: {matched} | unmatched: {len(diseases)-matched}")
