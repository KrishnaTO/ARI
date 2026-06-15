"""Ground ALL ARI core diseases to SNOMED using Gilda lexical matching.

Same Gilda method as the DOID grounder. Data source is LOCAL OMOP/Athena export:
data/2-databases/snomed/CONCEPT.csv + CONCEPT_SYNONYM.csv (vocabulary_id=SNOMED,
domain_id=Condition). No online sources.
"""
import csv, os, sys
from gilda import Grounder
from gilda.term import Term
from gilda.process import normalize
import openpyxl

csv.field_size_limit(sys.maxsize)
BASE = "/sessions/zen-keen-curie/mnt/ARI"
SN = f"{BASE}/data/2-databases/snomed"
CORE = f"{BASE}/data/4-reports/1_Core_ARI_Diseases.xlsx"
OUT = f"{BASE}/notebook/ari-grounding"

# 1) SNOMED Condition concepts: concept_id -> (snomed_code, name, invalid_reason)
concept = {}
with open(f"{SN}/CONCEPT.csv", encoding="utf-8") as f:
    r = csv.reader(f, delimiter="\t")
    next(r)
    for row in r:
        # concept_id, concept_name, domain_id, vocabulary_id, concept_class_id, standard_concept, concept_code, ...invalid_reason(9)
        if len(row) < 10: continue
        if row[3] == "SNOMED" and row[2] == "Condition" and row[5] == "S":
            concept[row[0]] = (row[6], row[1], row[9])
print("SNOMED Condition standard concepts:", len(concept))

terms = []
for cid, (code, name, inv) in concept.items():
    if name:
        terms.append(Term(normalize(name), name, "SNOMED", code, name, "name", "snomed", source_id=cid))

# 2) synonyms for those concepts
nsyn = 0
with open(f"{SN}/CONCEPT_SYNONYM.csv", encoding="utf-8") as f:
    r = csv.reader(f, delimiter="\t")
    next(r)
    for row in r:
        if len(row) < 2: continue
        cid = row[0]
        if cid in concept:
            syn = row[1]
            if syn:
                code, name, inv = concept[cid]
                terms.append(Term(normalize(syn), syn, "SNOMED", code, name, "synonym", "snomed", source_id=cid))
                nsyn += 1
print(f"terms: {len(terms)} (names + {nsyn} synonyms)")

grounder = Grounder(terms)

cws = openpyxl.load_workbook(CORE, read_only=True).active
hdr = [c for c in next(cws.iter_rows(min_row=1, max_row=1, values_only=True))]
i_ari, i_name, i_syn = hdr.index("ARI ID"), hdr.index("Preferred Name"), hdr.index("Synonyms")
i_existing = hdr.index("SNOMED Code(s)")
diseases = [(r[i_ari], r[i_name], r[i_syn] or "", r[i_existing] or "") for r in cws.iter_rows(min_row=2, values_only=True)]

def best_match(name, synonyms):
    candidates = [name] + [s.strip() for s in synonyms.split(";") if s.strip()]
    best = None; via = ""
    for txt in candidates:
        for m in grounder.ground(txt):
            if best is None or m.score > best.score:
                best = m; via = txt
        if best and best.score >= 0.9 and via == name:
            break
    return best, via

rows = []; matched = 0; agree = 0
for ari, name, syn, existing in diseases:
    m, via = best_match(name, syn)
    if m:
        matched += 1
        code = m.term.id; cid = m.term.source_id
        inv = concept.get(cid, ("","",""))[2]
        existing_codes = {c.strip() for c in existing.split(",") if c.strip()}
        match_existing = "Yes" if code in existing_codes else ("" if existing_codes else "n/a (no existing code)")
        if code in existing_codes: agree += 1
        rows.append([ari, name, code, m.term.entry_name, round(m.score,3), m.term.status,
                     via if via != name else "(preferred name)", cid,
                     "Obsolete" if inv else "", existing, match_existing])
    else:
        rows.append([ari, name, "", "", "", "", "no match", "", "", existing, ""])

with open(f"{OUT}/snomed_matches_all.csv", "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["ARI ID","Preferred Name","SNOMED Code","SNOMED Name","Score","Match Type",
                "Matched Via","OMOP ConceptID","SNOMED Status","Existing SNOMED (master)","Agrees w/ Existing"])
    w.writerows(rows)

print(f"Diseases: {len(diseases)} | SNOMED matched: {matched} | agrees w/ existing master code: {agree}")
