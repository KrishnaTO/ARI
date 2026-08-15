"""Build formatted xlsx reports from the Gilda grounding CSVs.
Outputs to data/4-reports: 6_DOID_Matches_All.xlsx, 7_SNOMED_Matches_All.xlsx."""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/sessions/zen-keen-curie/mnt/ARI"
HERE = f"{BASE}/notebook/ari-grounding"
OUTDIR = f"{BASE}/data/4-reports"
ATHENA_TERM = "https://athena.ohdsi.org/search-terms/terms/{}"
ATHENA_SEARCH = "https://athena.ohdsi.org/search-terms/terms?query={}"
DOID_PURL = "http://purl.obolibrary.org/obo/DOID_{}"

FONT="Arial"
HDR_FILL=PatternFill("solid",fgColor="1F3864")
HDR_FONT=Font(name=FONT,bold=True,color="FFFFFF",size=10)
CELL_FONT=Font(name=FONT,size=10)
LINK_FONT=Font(name=FONT,size=10,color="0563C1",underline="single")
WRAP=Alignment(vertical="top",wrap_text=True); TOP=Alignment(vertical="top")
thin=Side(style="thin",color="D9D9D9"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

def build(ws, headers, rows, widths, wrap_cols, link_col=None, link_fmt=None, fill_rules=None):
    ws.append(headers)
    for r in rows: ws.append(r)
    for c in range(1,len(headers)+1):
        cell=ws.cell(row=1,column=c); cell.fill=HDR_FILL; cell.font=HDR_FONT
        cell.alignment=Alignment(vertical="center",wrap_text=True); cell.border=BORDER
    ws.row_dimensions[1].height=28; ws.freeze_panes="A2"
    for ri in range(2,ws.max_row+1):
        for ci in range(1,len(headers)+1):
            cell=ws.cell(row=ri,column=ci); cell.font=CELL_FONT; cell.border=BORDER
            cell.alignment=WRAP if (ci-1) in wrap_cols else TOP
        if link_col is not None:
            val=ws.cell(row=ri,column=link_col).value
            if val:
                ws.cell(row=ri,column=link_col).hyperlink=link_fmt(val)
                ws.cell(row=ri,column=link_col).font=LINK_FONT
        if fill_rules: fill_rules(ws,ri)
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
    ws.auto_filter.ref=f"A1:{get_column_letter(len(headers))}1"

# ---- DOID report ----
d=list(csv.reader(open(f"{HERE}/doid_matches_all.csv")))
dh,drows=d[0],d[1:]
wb=Workbook(); ws=wb.active; ws.title="DOID Matches (All)"
build(ws, dh, drows, [13,40,14,34,8,14,26,24,12], wrap_cols={1,3,6},
      link_col=3, link_fmt=lambda v: DOID_PURL.format(v.split(":")[1]))

# Sheet 2 of report 6: full additional data on each MATCHED DOID disease
import json as _json
_recs=_json.load(open(f"{HERE}/doid_records.json"))
_idx={h:i for i,h in enumerate(dh)}
def _xref_group(xrefs,prefixes):
    out=[]
    for x in xrefs:
        for pre in prefixes:
            if x.upper().startswith(pre):
                out.append(x); break
    return "; ".join(out)
det_h=["ARI ID","Preferred Name","DOID","DOID Label","Definition",
       "Exact Synonyms","Related/Narrow Synonyms","SNOMED xrefs","MESH xrefs",
       "NCI xrefs","UMLS xrefs","ICD xrefs","Other xrefs","Deprecated"]
det=[]
for row in drows:
    doid=row[_idx["DOID"]]
    if not doid: continue
    r=_recs.get(doid,{})
    xr=r.get("xrefs",[])
    snomed=_xref_group(xr,["SNOMED"])
    mesh=_xref_group(xr,["MESH","MSH"])
    nci=_xref_group(xr,["NCI"])
    umls=_xref_group(xr,["UMLS"])
    icd=_xref_group(xr,["ICD10","ICD-10"])
    used=set(snomed.split("; ")+mesh.split("; ")+nci.split("; ")+umls.split("; ")+icd.split("; "))
    other="; ".join(x for x in xr if x not in used and not x.startswith("url:"))
    det.append([row[_idx["ARI ID"]],row[_idx["Preferred Name"]],doid,r.get("label",""),
                r.get("definition",""),"; ".join(r.get("exact_syn",[])),
                "; ".join(r.get("related_syn",[])+r.get("narrow_syn",[])),
                snomed,mesh,nci,umls,icd,other,"Yes" if r.get("deprecated") else ""])
ws_det=wb.create_sheet("Matched Disease Details")
build(ws_det, det_h, det, [13,34,12,30,70,40,40,22,20,16,22,18,30,11],
      wrap_cols={1,3,4,5,6,7,8,9,10,11,12},
      link_col=3, link_fmt=lambda v: DOID_PURL.format(v.split(":")[1]))

wb.save(f"{OUTDIR}/6_DOID_Matches_All.xlsx")

# ---- SNOMED report ----
s=list(csv.reader(open(f"{HERE}/snomed_matches_all.csv")))
sh,srows=s[0],s[1:]
i_agree=sh.index("Agrees w/ Existing")
AG=PatternFill("solid",fgColor="E2EFDA"); DIFF=PatternFill("solid",fgColor="FCE4D6")
def snomed_fill(ws,ri):
    v=ws.cell(row=ri,column=i_agree+1).value
    code=ws.cell(row=ri,column=3).value
    if v=="Yes": ws.cell(row=ri,column=i_agree+1).fill=AG
    elif code and v not in ("","n/a (no existing code)"): ws.cell(row=ri,column=i_agree+1).fill=DIFF
wb2=Workbook(); ws2=wb2.active; ws2.title="SNOMED Matches (All)"
i_omop=sh.index("OMOP ConceptID")
build(ws2, sh, srows, [13,38,14,34,8,12,24,14,12,24,16], wrap_cols={1,3,6,9},
      link_col=3, link_fmt=lambda v: ATHENA_SEARCH.format(v), fill_rules=snomed_fill)
for ri in range(2, ws2.max_row+1):
    ov=ws2.cell(row=ri,column=i_omop+1).value
    if ov:
        ws2.cell(row=ri,column=i_omop+1).hyperlink=ATHENA_TERM.format(ov); ws2.cell(row=ri,column=i_omop+1).font=LINK_FONT
wb2.save(f"{OUTDIR}/7_SNOMED_Matches_All.xlsx")
print("Wrote 6_DOID_Matches_All.xlsx and 7_SNOMED_Matches_All.xlsx")
print("DOID rows:",len(drows),"| SNOMED rows:",len(srows))

# ---- Report 5 (replace): DOID mapping for no-code diseases, Gilda-based ----
import openpyxl as _ox
core_ws=_ox.load_workbook(f"{BASE}/data/4-reports/1_Core_ARI_Diseases.xlsx",read_only=True).active
_h=[c for c in next(core_ws.iter_rows(min_row=1,max_row=1,values_only=True))]
_ia,_in,_isy,_ics=_h.index("ARI ID"),_h.index("Preferred Name"),_h.index("Synonyms"),_h.index("Code Status")
nocode={r[_ia]:(r[_in],r[_isy] or "",r[_ics]) for r in core_ws.iter_rows(min_row=2,values_only=True) if r[_ics]}
dmatch={r["ARI ID"]:r for r in csv.DictReader(open(f"{HERE}/doid_matches_all.csv"))}
r5=[]
for ari,(name,syn,cs) in nocode.items():
    m=dmatch[ari]
    note="Matched (Gilda lexical)" if m["DOID"] else "No DOID match (Gilda lexical)"
    r5.append([ari,name,syn,cs,m["DOID"],m["DOID Label"],m["Score"],m["Match Type"],
               m["Matched Via"],m["DOID SNOMED xref"],m["DOID Obsolete"],note])
wb5=Workbook(); ws5=wb5.active; ws5.title="DOID Mapping"
build(ws5,
      ["ARI ID","Preferred Name","Synonyms","Code Status","Matched DOID","DOID Label","Score",
       "Match Type","Matched Via","DOID SNOMED xref","DOID Obsolete","Match Note"],
      r5, [13,40,38,18,14,30,8,14,22,22,12,28], wrap_cols={1,2,5,8,11},
      link_col=5, link_fmt=lambda v: DOID_PURL.format(v.split(":")[1]))
wb5.save(f"{OUTDIR}/5_DOID_Mapping.xlsx")
print(f"Report 5 (DOID mapping, no-code, Gilda): {len(r5)} diseases | matched: {sum(1 for x in r5 if x[4])}")
