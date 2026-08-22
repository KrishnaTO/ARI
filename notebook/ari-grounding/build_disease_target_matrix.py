"""Build data/4-reports/8_Disease_Target_Mappings.xlsx.

One row per (disease, target database) pair for every core ARI disease and
every target database used by the ARI mapping set, plus one extra row wherever
a disease has more than one mapping into the same database. Mapping state comes
from mappings/ari.sssom.tsv; target labels come from
notebook/ari-grounding/target_labels.json (see resolve_target_labels.py); the
predicted match per pair comes from notebook/ari-grounding/target_predictions.json
(see predict_target_matches.py).

Run order: predict_target_matches.py -> resolve_target_labels.py -> this script.
"""
import csv
import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[2]
CORE = REPO / "data" / "4-reports" / "1_Core_ARI_Diseases.xlsx"
SSSOM = REPO / "mappings" / "ari.sssom.tsv"
LABELS = REPO / "notebook" / "ari-grounding" / "target_labels.json"
PREDICTIONS = REPO / "notebook" / "ari-grounding" / "target_predictions.json"
OUT = REPO / "data" / "4-reports" / "8_Disease_Target_Mappings.xlsx"

# Target databases of the ARI mapping set, with display name and IRI base
# (bases taken from the curie_map header of ari.sssom.tsv).
TARGETS = [
    ("SNOMEDCT", "SNOMED CT", "http://snomed.info/id/"),
    ("omop", "OMOP / Athena", "https://athena.ohdsi.org/search-terms/terms/"),
    ("DOID", "Human Disease Ontology", "http://purl.obolibrary.org/obo/DOID_"),
    ("MONDO", "Mondo Disease Ontology", "http://purl.obolibrary.org/obo/MONDO_"),
    ("ncit", "NCI Thesaurus", "http://purl.obolibrary.org/obo/NCIT_"),
    ("icd10cm", "ICD-10-CM", "http://purl.bioontology.org/ontology/ICD10CM/"),
    ("mesh", "MeSH", "http://id.nlm.nih.gov/mesh/"),
    ("umls", "UMLS", "https://uts.nlm.nih.gov/uts/umls/concept/"),
    ("ORPHA", "Orphanet", "https://www.orpha.net/en/disease/detail/"),
    ("OMIM", "OMIM", "https://omim.org/entry/"),
]

STATUS_CONFIRMED = "Confirmed match"
STATUS_REJECTED = "Rejected match"
STATUS_ABSENT = "No term in database"
STATUS_UNREVIEWED = "Not reviewed"

VERDICT_AGREES = "Agrees with curated"
VERDICT_DIFFERS = "Differs from curated"
VERDICT_CANDIDATE = "Candidate for review"
VERDICT_CONTRADICTS = "Contradicts no-term finding"

# --- inputs ------------------------------------------------------------------
wb_core = openpyxl.load_workbook(CORE, read_only=True)
ws_core = wb_core["Core ARI Diseases"]
head = [c.value for c in next(ws_core.iter_rows(min_row=1, max_row=1))]
i_id = head.index("ARI ID")
i_name = head.index("Preferred Name")
i_syn = head.index("Synonyms")
diseases = []            # (ari_id, name, synonyms, in_core)
for rec in ws_core.iter_rows(min_row=2, values_only=True):
    if not rec[i_id]:
        continue
    syn = rec[i_syn]
    syn = "" if syn in (None, "None") else str(syn)
    diseases.append((rec[i_id], rec[i_name], syn, True))
wb_core.close()

lines = [l for l in SSSOM.open(encoding="utf-8") if not l.startswith("#")]
mappings = list(csv.DictReader(lines, delimiter="\t"))
labels = json.loads(LABELS.read_text(encoding="utf-8"))

predictions = json.loads(PREDICTIONS.read_text(encoding="utf-8"))

by_pair = defaultdict(list)          # (ari_id, prefix) -> [mapping rows]
for m in mappings:
    by_pair[(m["subject_id"], m["object_source"])].append(m)

confirmed_ids = defaultdict(set)     # (ari_id, prefix) -> {curated object id}
for m in mappings:
    if m["predicate_modifier"] != "Not" and not m["object_id"].startswith("sssom:"):
        confirmed_ids[(m["subject_id"], m["object_source"])].add(m["object_id"])

# Mapping subjects that are not in the core disease list are still reported,
# flagged so the gap is visible rather than silently dropped.
known = {d[0] for d in diseases}
extra = {}
for m in mappings:
    if m["subject_id"] not in known:
        extra.setdefault(m["subject_id"], m["subject_label"])
for ari_id, label in sorted(extra.items()):
    diseases.append((ari_id, label, "", False))


def classify(m):
    """(status, mapping type) for one SSSOM row."""
    if m["object_id"].startswith("sssom:NoTermFound"):
        return STATUS_ABSENT, "manual-absent"
    if m["predicate_modifier"] == "Not":
        return STATUS_REJECTED, "manual-negative"
    return STATUS_CONFIRMED, "manual"


COLUMNS = [
    ("ARI ID", 14), ("Disease Name", 42), ("Synonyms", 52), ("In Core List", 12),
    ("Target Database", 16), ("Target Database Name", 24),
    ("Mapping Status", 20), ("Mapping Type", 16),
    ("Mapping Predicate", 17), ("Predicate Modifier", 17), ("Mapping Justification", 28),
    ("Target Mapping ID", 20), ("Target Mapping Name", 46),
    ("Target Mapping Name Source", 30), ("Target URL", 46),
    ("Curator", 20), ("Mapping Date", 13),
    ("Predicted Mapping ID", 20), ("Predicted Mapping Name", 46),
    ("Prediction Method", 24), ("Prediction Support", 17),
    ("Prediction Evidence", 62), ("Prediction vs Curated", 21),
    ("Other Predicted IDs", 34), ("Predicted URL", 46),
]


def prediction_for(ari_id, prefix, base, status):
    """Prediction columns for one (disease, database) pair."""
    cands = predictions.get("%s|%s" % (ari_id, prefix))
    if not cands:
        return ["", "", "", "", "", "", "", ""]
    top = cands[0]
    curie = "%s:%s" % (prefix, top["id"])
    label, _ = labels.get(curie, [top["name"], ""])
    if status == STATUS_CONFIRMED:
        verdict = (VERDICT_AGREES if curie in confirmed_ids[(ari_id, prefix)]
                   else VERDICT_DIFFERS)
    elif status == STATUS_ABSENT:
        verdict = VERDICT_CONTRADICTS
    else:
        verdict = VERDICT_CANDIDATE
    others = ", ".join("%s:%s" % (prefix, c["id"]) for c in cands[1:])
    return [curie, label or top["name"], top["method"], top["support"],
            top["evidence"], verdict, others, base + top["id"]]


rows = []
for ari_id, name, syn, in_core in diseases:
    for prefix, db_name, base in TARGETS:
        found = by_pair.get((ari_id, prefix), [])
        in_core_flag = "Yes" if in_core else "No"
        if not found:
            rows.append([ari_id, name, syn, in_core_flag, prefix, db_name,
                         STATUS_UNREVIEWED, "", "", "", "", "", "", "", "", "", ""]
                        + prediction_for(ari_id, prefix, base, STATUS_UNREVIEWED))
            continue
        for m in found:
            status, mtype = classify(m)
            oid = m["object_id"]
            if status == STATUS_ABSENT:
                target_id = target_label = target_src = url = ""
            else:
                target_id = oid
                target_label, target_src = labels.get(oid, ["", ""])
                url = base + oid.split(":", 1)[1]
            rows.append([ari_id, name, syn, in_core_flag, prefix, db_name,
                         status, mtype, m["predicate_id"], m["predicate_modifier"],
                         m["mapping_justification"], target_id, target_label,
                         target_src, url, m["author_id"], m["mapping_date"]]
                        + prediction_for(ari_id, prefix, base, status))

# --- workbook ----------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
STATUS_FILL = {
    STATUS_CONFIRMED: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_REJECTED: PatternFill("solid", fgColor="FFC7CE"),
    STATUS_ABSENT: PatternFill("solid", fgColor="FFEB9C"),
    STATUS_UNREVIEWED: PatternFill("solid", fgColor="F2F2F2"),
}
VERDICT_FILL = {
    VERDICT_AGREES: PatternFill("solid", fgColor="C6EFCE"),
    VERDICT_DIFFERS: PatternFill("solid", fgColor="FFD8A8"),
    VERDICT_CANDIDATE: PatternFill("solid", fgColor="DDEBF7"),
    VERDICT_CONTRADICTS: PatternFill("solid", fgColor="FFEB9C"),
}


def style_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Disease-Target Mappings"
ws.append([c[0] for c in COLUMNS])
style_header(ws)
for r in rows:
    ws.append(r)

names = [c[0] for c in COLUMNS]
i_status = names.index("Mapping Status")
i_verdict = names.index("Prediction vs Curated")
link_cols = [names.index("Target URL"), names.index("Predicted URL")]
for row in ws.iter_rows(min_row=2):
    row[i_status].fill = STATUS_FILL[row[i_status].value]
    if row[i_verdict].value:
        row[i_verdict].fill = VERDICT_FILL[row[i_verdict].value]
    for i in link_cols:
        url = row[i].value
        if url:
            row[i].hyperlink = url
            row[i].font = Font(color="0563C1", underline="single")
for idx, (_, width) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = width
ws.freeze_panes = "C2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLUMNS)), ws.max_row)

# --- summary -----------------------------------------------------------------
ws2 = wb.create_sheet("Summary")
ws2.append(["Target Database", "Target Database Name", "Diseases with confirmed match",
            "Confirmed mappings", "Rejected mappings", "No term in database",
            "Diseases not reviewed", "Unreviewed with a prediction",
            "Prediction agrees with curated", "Prediction differs from curated"])
style_header(ws2)
i_pred = names.index("Predicted Mapping ID")
for prefix, db_name, _ in TARGETS:
    sub = [r for r in rows if r[4] == prefix]
    conf = [r for r in sub if r[i_status] == STATUS_CONFIRMED]
    unreviewed = [r for r in sub if r[i_status] == STATUS_UNREVIEWED]
    ws2.append([prefix, db_name, len({r[0] for r in conf}), len(conf),
                sum(1 for r in sub if r[i_status] == STATUS_REJECTED),
                sum(1 for r in sub if r[i_status] == STATUS_ABSENT),
                len(unreviewed),
                sum(1 for r in unreviewed if r[i_pred]),
                sum(1 for r in conf if r[i_verdict] == VERDICT_AGREES),
                sum(1 for r in conf if r[i_verdict] == VERDICT_DIFFERS)])
ws2.append([])
ws2.append(["Diseases", len(diseases)])
ws2.append(["Core diseases (1_Core_ARI_Diseases.xlsx)", sum(1 for d in diseases if d[3])])
ws2.append(["Mapping subjects not in core list", sum(1 for d in diseases if not d[3])])
ws2.append(["Rows", len(rows)])
ws2.append(["Mappings in ari.sssom.tsv", len(mappings)])
ws2.append(["Rows carrying a prediction", sum(1 for r in rows if r[i_pred])])
for idx, width in enumerate([18, 26, 26, 20, 20, 22, 22, 26, 26, 26], start=1):
    ws2.column_dimensions[get_column_letter(idx)].width = width

# --- legend ------------------------------------------------------------------
ws3 = wb.create_sheet("Legend")
ws3.append(["Mapping Status", "Meaning", "Source in ari.sssom.tsv"])
style_header(ws3)
for status, meaning, source in [
    (STATUS_CONFIRMED, "Curator confirmed this disease equals this target term",
     "skos:exactMatch, no predicate_modifier"),
    (STATUS_REJECTED, "Curator reviewed this target term and rejected it as a match",
     "skos:exactMatch, predicate_modifier = Not"),
    (STATUS_ABSENT, "Curator searched the database and found no matching term",
     "object_id = sssom:NoTermFound"),
    (STATUS_UNREVIEWED, "No mapping row exists for this disease and database",
     "absent from the mapping set"),
]:
    ws3.append([status, meaning, source])
    ws3.cell(ws3.max_row, 1).fill = STATUS_FILL[status]
ws3.append([])
ws3.append(["Prediction Method", "Meaning", "Ranking"])
for cell in ws3[ws3.max_row]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEAD_FILL
for method, meaning, rank in [
    ("xref via curated anchor",
     "A Mondo or DOID term reached from one of the disease's curated mappings "
     "cross-references this target term",
     "Prediction Support = weighted count of the disease identifiers reaching that hub term"),
    ("lexical grounding",
     "Gilda lexical match of the disease name or a synonym against the target vocabulary",
     "From 6_DOID_Matches_All.xlsx / 7_SNOMED_Matches_All.xlsx"),
    ("xref via lexical anchor",
     "As above, but the hub term was reached from a lexical match rather than a curated mapping",
     "Weakest route; review before accepting"),
]:
    ws3.append([method, meaning, rank])
ws3.append([])
ws3.append(["Prediction vs Curated", "Meaning", ""])
for cell in ws3[ws3.max_row]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEAD_FILL
for verdict, meaning in [
    (VERDICT_AGREES, "The prediction reproduces the curated mapping (validation)"),
    (VERDICT_DIFFERS, "The prediction names a different term than the curated mapping - review"),
    (VERDICT_CANDIDATE, "No confirmed mapping yet; the prediction is a candidate to review"),
    (VERDICT_CONTRADICTS,
     "A curator recorded no term in this database, but a prediction was still found - "
     "usually a broader or neighbouring concept, so check before acting"),
]:
    ws3.append([verdict, meaning])
    ws3.cell(ws3.max_row, 1).fill = VERDICT_FILL[verdict]
ws3.append([])
ws3.append(["Predictions never include a term the curators rejected for that disease and "
            "database, and predicted SNOMED codes are restricted to standard, non-retired "
            "concepts."])
ws3.append([])
ws3.append(["Target Mapping Name sources"])
ws3.cell(ws3.max_row, 1).font = Font(bold=True)
for line in [
    "SNOMED CT, OMOP, ICD-10-CM, MeSH: data/2-databases OMOP vocabulary exports (CONCEPT.csv)",
    "Human Disease Ontology: data/2-databases/doid.owl",
    "Mondo: data/2-databases/mondo.obo",
    "OMIM: data/2-databases/OMIM.ttl",
    "Orphanet, NCI Thesaurus: EBI OLS4 API (no usable local copy)",
    "UMLS: label of the DOID or Mondo term cross-referencing the CUI (no local UMLS label source)",
]:
    ws3.append([line])
for idx, width in enumerate([24, 62, 46], start=1):
    ws3.column_dimensions[get_column_letter(idx)].width = width

wb.save(OUT)
print("wrote %s (%d rows, %d diseases)" % (OUT, len(rows), len(diseases)))
